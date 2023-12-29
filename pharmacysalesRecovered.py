import sys
from selenium.webdriver.support.select import Select
from selenium.webdriver.edge.webdriver import WebDriver
from selenium.webdriver.edge import webdriver
from pygetwindow._pygetwindow_win import NULL
#from psutil import msg
import psutil
sys.path.append('..\..\libraries\standard') 
sys.path.append('../../libraries/standard')
sys.path.append('..\..\libraries\Application_specific')
sys.path.append('..\..\libraries\Workflow_Specific')
import admin
from admin import FromConfigFile
import WF_Pharmacy
from WF_Pharmacy import InPharmacycommon
from WF_Pharmacy import InSalesPharmacy
#import WF_FrontOffice
#from WF_FrontOffice import InNewOpRegistration
from datetime import datetime, timedelta, date
import datetime
import time
import pyautogui
import openpyxl
from openpyxl import load_workbook
import common_reader
from common_reader import Capturing
from Tix import COLUMN
from PIL.TiffImagePlugin import MM
import common_importstatements
from common_importstatements import *
import psutil
from selenium.common.exceptions import StaleElementReferenceException  
from selenium.common.exceptions import ElementClickInterceptedException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import UnexpectedAlertPresentException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select


class salespharmacy(Selenium2Library):
    dict = admin.FromConfigFile.dict
    objects = common_reader.Capturing.objects
    d = Capturing().data_off("sales_pharm")
    def multiple_pharmacysales(self):
        self.set_selenium_implicit_wait(30)
        print datetime.datetime.now()
        admin.FromConfigFile().driving_browser_and_url()
        admin.FromConfigFile().logging("pharmsales")
        self._cache.current = self.dict['BROWSER']
        self.browser = self._current_browser()
        WF_Pharmacy.InPharmacycommon().selecting_the_frame()
        WF_Pharmacy.InPharmacycommon().selecting_storename()
        admin.FromConfigFile().loading_menu_of_link("//Pharmacy/tIPMedIssue.aspx?opt=2")
        WF_Pharmacy.InPharmacycommon().selecting_the_frame()
        WF_Pharmacy.InSalesPharmacy().selecting_the_frame()
        wb = load_workbook('D:\workspace\Automate_BB15_WF\datas\Bb_datas.xlsx')
        ws = wb["sales_pharm"]
        num_rows = ws.max_row
        self.r = 1
        i = 1
        j = 0
        salestobedone = 1000000
        while self.r < num_rows and j < salestobedone:
         self.r = self.r + 1
         statusvalue = ws.cell(row=(self.r),column=2).value
         print statusvalue
         if statusvalue == "yes":
            rwno = self.r - 1
            start_time = datetime.datetime.now()
            if self.d[rwno]["salespharm_regno"] is not None:
               WF_Pharmacy.InSalesPharmacy().entering_regno_with_data(rwno)
               print "reg no:    ", self.d[rwno]["salespharm_regno"]
               if self.d[rwno]["salespharm_item"] is not None:
                  WF_Pharmacy.InSalesPharmacy().entering_item_with_data(rwno)
                  if self.d[rwno]["salespharm_batchname"] is not None:
                     WF_Pharmacy.InSalesPharmacy().entering_batchname_with_data(rwno)
                     if self.d[rwno]["salespharm_issqty"] is not None:
                        WF_Pharmacy.InSalesPharmacy().entering_issqty_with_data(rwno)
                        print "salespharm_transactionflag", self.d[rwno]["salespharm_transactionflag"]
                        print "salespharm_paymentflag", self.d[rwno]["salespharm_paymentflag"]
                        if (self.d[rwno]["salespharm_transactionflag"] is not None) and (self.d[rwno]["salespharm_paymentflag"] is not None):
                           print "inside payment options"
                           self.wait_until_element_is_visible(self.objects["Pharm_Sales_addstkbtn"], 15, "add button was not visible")
                           self.click_button(self.objects["Pharm_Sales_addstkbtn"])
                           j,resmsg = WF_Pharmacy.InSalesPharmacy().selecting_paymenttype(rwno,wb,j)
                        else:
                            print "inside else payment option"
                            ws.cell(row=(self.r),column=3).value = "Improper payment option"
                            j = j + 1
                            self.wait_until_element_is_visible(self.objects["Pharm_Sales_cancelbtn"], 15, "cancel button was not visible")
                            self.click_button(self.objects["Pharm_Sales_cancelbtn"])
                     else:
                       ws.cell(row=(self.r),column=3).value = "Improper issue quantity"
                       j = j + 1
                       self.wait_until_element_is_visible(self.objects["Pharm_Sales_cancelbtn"], 15, "cancel button was not visible")
                       self.click_button(self.objects["Pharm_Sales_cancelbtn"])
                  else:
                     ws.cell(row=(self.r),column=3).value = "Improper batch name"
                     j = j + 1
                     self.wait_until_element_is_visible(self.objects["Pharm_Sales_cancelbtn"], 15, "cancel button was not visible")
                     self.click_button(self.objects["Pharm_Sales_cancelbtn"])
               else:
                  ws.cell(row=(self.r),column=3).value = "Improper item name"
                  j = j + 1
                  self.wait_until_element_is_visible(self.objects["Pharm_Sales_cancelbtn"], 15, "cancel button was not visible")
                  self.click_button(self.objects["Pharm_Sales_cancelbtn"])
            else:
                ws.cell(row=(self.r),column=3).value = "Improper registration number"
                j = j + 1
                self.wait_until_element_is_visible(self.objects["Pharm_Sales_cancelbtn"], 15, "cancel button was not visible")
                self.click_button(self.objects["Pharm_Sales_cancelbtn"])
            try:
                j = WF_Pharmacy.InSalesPharmacy().selecting_save_btn(wb,j)
                self.wait_until_element_is_visible(self.objects["Pharm_Sales_cancelbtn"], 15, "cancel button was not visible")
                self.click_button(self.objects["Pharm_Sales_cancelbtn"])
                self._handle_alert(True)
            except:
                if resmsg == "Improper payment option or payment transaction option are given":
                    resmsg = ws.cell(row=(self.r),column=3).value = "Improper payment option or payment transaction option are given"
                else:
                    print "save or cancel button issue"
                    ws.cell(row=(self.r),column=3).value = "problem during save or cancel button"
                    j = j + 1
            else:
                print "saved data"
                ws.cell(row=(self.r),column=3).value = resmsg
                j = j + 1
            end_time = datetime.datetime.now()
            diff_time = end_time-start_time
            diff_seconds =  diff_time.total_seconds()
            ws = wb["sales_pharm"]
            ws = wb.active
            ws.cell(row=(self.r),column=4).value = diff_seconds
            ws = wb["sales_pharm"]
            rampercentage =  psutil.virtual_memory()[2]
            ws.cell(row=(self.r),column=6).value = rampercentage
            if diff_seconds > 50 or rampercentage > 80:
               self.unselect_frame()
               admin.FromConfigFile().Logoff()
               ws = wb["sales_pharm"]
               ws.cell(row=(self.r),column=5).value = "Crossed 50 secs or ram percent crossed 80 sec, so re-login browser"
               admin.FromConfigFile().Test_Teardown()
               admin.FromConfigFile().driving_browser_and_url()
               admin.FromConfigFile().logging("pharmsales")
               self._cache.current = self.dict['BROWSER']
               self.browser = self._current_browser()
               WF_Pharmacy.InPharmacycommon().selecting_the_frame()
               WF_Pharmacy.InPharmacycommon().selecting_storename()
               admin.FromConfigFile().loading_menu_of_link("//Pharmacy/tIPMedIssue.aspx?opt=2")
               WF_Pharmacy.InSalesPharmacy().selecting_the_frame()
        wb.save('D:\workspace\Automate_BB15_WF\datas\Bb_datas.xlsx')
        wb.close()
        WF_Pharmacy.InSalesPharmacy().unselecting_the_frame()
        #admin.FromConfigFile().Logoff()
        self.close_browser()
        #admin.FromConfigFile().Test_Teardown()

salespharmacy().multiple_pharmacysales()